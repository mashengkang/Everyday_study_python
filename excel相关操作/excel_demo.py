import datetime

from openpyxl import Workbook, load_workbook


# # 实例化
# wb = Workbook()
# # 获取当前active的sheet
# sheet = wb.active
# print(sheet.title)  # 打印当前sheet名
# sheet.title = 'sheet学习01'  # 更改sheet名
#
# # 写入数据
# # 方法一：
# sheet['B9'] = 'yaya'
# sheet['C9'] = 'mama'
# # 方法二"append,会将数据添加在最后一行
# sheet.append([1,2,3])
# # 方法三：python类型会被自动转换
# sheet['A3'] = datetime.datetime.now().strftime("%Y-%m-%d")
#
#
# # 保存
# wb.save('excel_test.xlsx')



wb = load_workbook('excel_test.xlsx')

print(wb.sheetnames)  # 推荐使用这种方式获取sheetname
print(wb.get_sheet_names())  # python推荐使用第一种
sheet = wb['sheet学习01']  # 推荐使用这种方式操作某个sheet
sheet = wb.get_sheet_by_name('sheet学习01')  # python推荐使用第一种
print(sheet['C9'])  # <Cell 'sheet学习01'.C9>
print(sheet['C9'].value)  # 获取某个单元格值：mama

for cell in sheet['A4:A9']:
    print(cell)  # (<Cell 'sheet学习01'.A9>,)是一个元组
    print(cell[0].value)  # <Cell 'sheet学习01'.A9> 所对应的值

# 按行遍历
for row in sheet:
    print(row)
    for cell in row:
        print(cell.value)
# 其中 row是一行中所有的数据，元组形式：
# row：(<Cell 'sheet学习01'.A12>, <Cell 'sheet学习01'.B12>, <Cell 'sheet学习01'.C12>, <Cell 'sheet学习01'.D12>)
# 遍历的顺序是：A1,A2...B1,B2,,
# 空的单元格显示None

# 按列遍历
for column in sheet.columns:
    for cell in column:
        print(cell.value,end=',')
    print()

# 指定行&列进行遍历
# 遍历6-11行，前3列的数据
for row in sheet.iter_rows(min_row=6, max_row=11,max_col=3):
    for cell in row:
        print(cell.value,end=',')
    print()
# 遍历2-4列，第3行到第6行的数据
for col in sheet.iter_cols(min_col=2, max_col=4,max_row=6,min_row=3):
    for cell in col:
        print(cell.value,end=',')
    print()

# 删除工作表
# 方法1
wb.remove(sheet)
# 方法2
del wb[sheet]
